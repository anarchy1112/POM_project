from openpyxl import load_workbook



def dataGet(sheetname):
    wb=load_workbook(r"C:\Users\AZ\POM_project\excel\info.xlsx")
    sheet=wb[sheetname]
    totRow=sheet.max_row
    totCol=sheet.max_column
    main=[]
    for i in range(2,totRow+1):
        temp=[]
        for j in range(1, totCol+1):
            data=sheet.cell(row=i, column=j).value
            temp.insert(j,data)
        main.insert(i,temp)
    return main

