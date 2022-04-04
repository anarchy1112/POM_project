from openpyxl import load_workbook

def getRowCount(path,sheetName):
    wb=load_workbook(path, sheetName)
    sheet=wb[sheetName]
    return sheet.max_row


def getColCount(path,sheetName):
    wb = load_workbook(path, sheetName)
    sheet = wb[sheetName]
    return sheet.max_column

def getCellValue(path,sheetName,row_num, col_num):
    wb = load_workbook(path, sheetName)
    sheet = wb[sheetName]
    return sheet.cell(row=row_num, col=col_num).value

def setCellValue(path,sheetName,row_num, col_num, data):
    wb = load_workbook(path, sheetName)
    sheet = wb[sheetName]
    sheet.cell(row=row_num, col=col_num).value=data
    wb.save()